import openpyxl
import json
from rich.console import Console
console = Console()


def read_design(input, logfile) -> json:
    """
    Summary: Function that accepts a design xlsx file as input and converts in into a dictionary

    Args:
        input (xlsx): A microsoft excel file that contains the network design

    Returns:
        json: A dictionary of the contents in each tab of the excel file.
            It uses the column headings as the key.
    """
    # variable for end result which is a dictionary of dictionaries
    design_facts = {"facts": {}}
    # working variable (temp)
    spreadsheet = {}

    try:
        # open input file
        wb = openpyxl.load_workbook(input)
        # loop through each sheet
        for sheet in wb.sheetnames:
            # set current sheet
            current_sheet = wb[sheet]
            # variable to read the contents of current sheet
            spreadsheet[sheet] = []
            # variable to capture the dictionary keys
            dict_keys = []
            # loop through the contents of the first value in each column and assign as dict keys
            for c in range(1, current_sheet.max_column + 1):
                dict_keys.append(current_sheet.cell(row=1, column=c).value)
            # loop through the contents of the second row to the last row and assign value to keys
            for r in range(2, current_sheet.max_row + 1):
                temp_dict = {}
                # loop through each column in the row, read the value and assign to value
                for c in range(1, current_sheet.max_column + 1):
                    value = str(current_sheet.cell(row=r, column=c).value)
                    # # if value is None, then replace with empty string
                    if value == "None":
                        None
                    #     # value = ""
                    #     temp_dict[dict_keys[c - 1]] = value
                    # if semi-colon detected in value convert it to a list
                    elif value.find(";") != -1:
                        value = value.split(";")
                        # temporary value to remove any white spaces before and after
                        clean_value = []
                        for val in value:
                            clean_value.append(val.strip())
                        temp_dict[dict_keys[c - 1]] = clean_value
                    # if comma detected in value convert it to a list
                    elif value.find(",") != -1:
                        value = value.split(",")
                        # temporary value to remove any white spaces before and after
                        clean_value = []
                        for val in value:
                            clean_value.append(val.strip())
                        temp_dict[dict_keys[c - 1]] = clean_value
                    # else assign value to variable
                    else:
                        temp_dict[dict_keys[c - 1]] = value.strip()
                spreadsheet[sheet].append(temp_dict)

    except IOError:
        print("unable to read excel file")

    design_facts["facts"] = spreadsheet
    # print (json.dumps(design_facts, indent=4))
    console.print(f"[bold green] Reading design file {input}[/bold green]")
    logfile.write(f"Reading design file {input}\n")
    return design_facts
